from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = 'Hello'
ws['B1'] = 'World'
wb.save('simple_test.xlsx')
print("成功创建simple_test.xlsx")