from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "功能测试"

# 1. 添加基础数据
sheet['A1'] = '项目'
sheet['B1'] = '金额'
sheet['C1'] = '增长率'
sheet['D1'] = '备注'

# 添加数据行
data = [
    ['产品A', 1000, 0.15, '第一季度'],
    ['产品B', 2000, 0.20, '第一季度'],
    ['产品C', 1500, 0.10, '第一季度'],
    ['产品D', 3000, 0.25, '第一季度']
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# 2. 添加公式
sheet['B6'] = '=SUM(B2:B5)'
sheet['B7'] = '=AVERAGE(B2:B5)'
sheet['B8'] = '=MAX(B2:B5)'
sheet['B9'] = '=MIN(B2:B5)'

# 添加标签
sheet['A6'] = '总计'
sheet['A7'] = '平均值'
sheet['A8'] = '最大值'
sheet['A9'] = '最小值'

# 3. 添加增长率计算公式
sheet['C6'] = '=AVERAGE(C2:C5)'

# 4. 添加格式化
# 标题行格式
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill('solid', fgColor='4472C4')
header_alignment = Alignment(horizontal='center', vertical='center')

for col in range(1, 5):
    cell = sheet.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 数据格式
data_font = Font(size=11)
for row in range(2, 10):
    for col in range(1, 5):
        cell = sheet.cell(row=row, column=col)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

# 数字格式
for row in range(2, 6):
    sheet.cell(row=row, column=2).number_format = '#,##0'
    sheet.cell(row=row, column=3).number_format = '0.0%'

# 总计行格式
total_fill = PatternFill('solid', fgColor='D9E2F3')
for row in range(6, 10):
    for col in range(1, 5):
        cell = sheet.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = Font(bold=True, size=11)

# 5. 调整列宽
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12
sheet.column_dimensions['D'].width = 15

# 6. 添加边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in range(1, 10):
    for col in range(1, 5):
        cell = sheet.cell(row=row, column=col)
        cell.border = thin_border

# 7. 添加第二个工作表
sheet2 = wb.create_sheet("数据分析")
sheet2['A1'] = '数据分析工作表'
sheet2['A1'].font = Font(bold=True, size=14, color='4472C4')

# 从第一个工作表引用数据
sheet2['A3'] = '来自功能测试的数据：'
sheet2['A4'] = '=功能测试!B6'
sheet2['A5'] = '=功能测试!B7'

# 8. 保存文件
wb.save('功能测试.xlsx')
print("Excel文件创建成功：功能测试.xlsx")