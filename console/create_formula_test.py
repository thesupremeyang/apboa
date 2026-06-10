from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "公式测试"

# 添加数据
ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

# 添加公式
ws['B1'] = '=SUM(A1:A3)'
ws['B2'] = '=AVERAGE(A1:A3)'
ws['B3'] = '=MAX(A1:A3)'

# 添加标签
ws['C1'] = '总和'
ws['C2'] = '平均值'
ws['C3'] = '最大值'

wb.save('formula_test.xlsx')
print("公式测试文件创建成功：formula_test.xlsx")